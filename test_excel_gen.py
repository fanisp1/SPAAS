import sys
sys.path.insert(0, 'C:/SPAAS/backend')

from app.hypercube import hypercube_suppress, ProtectionRules
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import numpy as np

# Load test data
df = pd.read_excel('C:/SPAAS/realistic_test_data.xlsx')
print(f'Loaded data: {df.shape}')

# Run suppression
protection_rules = ProtectionRules(min_frequency=10)
suppressed_data, statistics = hypercube_suppress(df, protection_rules)
print(f'Suppressions: {len(statistics.get("suppressed_cells", []))}')

# Get coords
suppressed_coords = set()
for cell in statistics.get('suppressed_cells', []):
    suppressed_coords.add((cell['row'], cell['col']))

# Create Excel
wb = Workbook()
ws = wb.active

# Write header
for c_idx, col_name in enumerate(suppressed_data.columns, 1):
    ws.cell(row=1, column=c_idx, value=str(col_name))

# Write data
for r_idx, (_, row_data) in enumerate(suppressed_data.iterrows(), 2):
    for c_idx, col_name in enumerate(suppressed_data.columns, 1):
        value = row_data[col_name]
        
        if pd.isna(value):
            cell_value = None
        elif isinstance(value, (int, float, np.number)):
            cell_value = float(value) if not np.isnan(value) else None
        else:
            cell_value = str(value)
        
        cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
        
        data_row = r_idx - 2
        data_col = c_idx - 1
        if (data_row, data_col) in suppressed_coords:
            cell.font = Font(color='FF0000', bold=True)

# Save
stream = BytesIO()
wb.save(stream)
print(f'Excel size: {len(stream.getvalue())} bytes')

# Save to file for testing
with open('C:/SPAAS/test_direct.xlsx', 'wb') as f:
    f.write(stream.getvalue())
print('Saved to test_direct.xlsx - try opening this file')
