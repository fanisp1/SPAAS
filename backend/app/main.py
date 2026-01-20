from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import numpy as np
from io import StringIO, BytesIO
import json
from pathlib import Path
import tempfile
import os

from .hypercube import hypercube_suppress, ProtectionRules
from .batch_parser import parse_batch_file, BatchFile
from .tauargus_formats import TauArgusFormatHandler
from .primary_suppression import (
    apply_primary_suppression_to_file,
    ProtectionRules as PrimaryProtectionRules,
    PrimarySuppressionEngine
)

app = FastAPI(
    title="SPAAS Modernized API",
    description="Statistical Package for Automated Anonymization Software - Modern Python Implementation",
    version="1.0.0"
)

# Configure CORS - Allow all origins for development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins in development
    allow_credentials=False,  # Must be False when allow_origins is *
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

@app.get("/")
def read_root():
    return {"message": "SPAAS Modernized Backend is running."}

@app.post("/analyze/")
async def analyze_table(file: UploadFile = File(...)):
    """Analyze uploaded table and return basic statistics"""
    contents = await file.read()
    filename = file.filename.lower()

    # Decide how to load based on file extension
    if filename.endswith('.csv'):
        # CSV: decode bytes to string
        s = str(contents, 'utf-8')
        df = pd.read_csv(StringIO(s))
    elif filename.endswith('.xlsx'):
        # Excel: read directly from bytes
        df = pd.read_excel(BytesIO(contents))
    else:
        raise HTTPException(status_code=400, detail="File must be a CSV or Excel (.xlsx) file")

    # Solution: Replace NaN with None for JSON compatibility
    sample_head = df.head().replace({np.nan: None}).to_dict(orient="records")
    
    info = {
        "rows": df.shape[0],
        "columns": df.shape[1],
        "column_names": df.columns.tolist(),
        "sample_head": sample_head
    }
    return info


@app.post("/suppress/hypercube/")
async def hypercube_suppression(
    file: UploadFile = File(...),
    min_frequency: int = 3,
    dominance_n: int = 1,
    dominance_k: float = 80.0,
    p_percent: float = 10.0
):
    """
    Apply hypercube secondary suppression to uploaded table
    
    Parameters:
    - file: CSV or Excel file containing the data table
    - min_frequency: Minimum frequency threshold (default: 3)
    - dominance_n: Number of top contributors for dominance rule (default: 1)
    - dominance_k: Percentage threshold for dominance (default: 80.0)
    - p_percent: P-percent rule threshold (default: 10.0)
    
    Returns:
    - JSON with suppressed table and statistics
    """
    try:
        # Read uploaded file
        contents = await file.read()
        filename = file.filename.lower()
        
        if filename.endswith('.csv'):
            s = str(contents, 'utf-8')
            df = pd.read_csv(StringIO(s))
        elif filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(contents))
        else:
            raise HTTPException(
                status_code=400,
                detail="File must be a CSV or Excel (.xlsx) file"
            )
        
        # Create protection rules from parameters
        protection_rules = ProtectionRules(
            min_frequency=min_frequency,
            dominance_n=dominance_n,
            dominance_k=dominance_k,
            p_percent=p_percent
        )
        
        # Apply hypercube suppression
        suppressed_data, statistics = hypercube_suppress(
            data=df,
            protection_rules=protection_rules
        )
        
        # Convert suppressed data to JSON-compatible format
        suppressed_json = suppressed_data.replace({np.nan: None}).to_dict(orient="records")
        
        result = {
            "status": "success",
            "method": "hypercube",
            "statistics": statistics,
            "suppressed_data": suppressed_json,
            "column_names": suppressed_data.columns.tolist()
        }
        
        return result
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error processing table: {str(e)}"
        )


@app.post("/suppress/hypercube/download/")
async def hypercube_suppression_download(
    file: UploadFile = File(...),
    min_frequency: int = Form(3),
    dominance_n: int = Form(1),
    dominance_k: float = Form(80.0),
    p_percent: float = Form(10.0),
    output_format: str = Form("csv")
):
    """
    Apply hypercube suppression and download result as file with red highlighting
    
    Parameters:
    - file: CSV or Excel file containing the data table
    - min_frequency: Minimum frequency threshold (default: 3)
    - dominance_n: Number of top contributors for dominance rule (default: 1)
    - dominance_k: Percentage threshold for dominance (default: 80.0)
    - p_percent: P-percent rule threshold (default: 10.0)
    - output_format: Output format ('csv' or 'excel')
    
    Returns:
    - Downloadable file with suppressed data (red cells in Excel)
    """
    try:
        # Read uploaded file
        contents = await file.read()
        filename = file.filename.lower()
        
        if filename.endswith('.csv'):
            s = str(contents, 'utf-8')
            df = pd.read_csv(StringIO(s))
        elif filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(contents))
        else:
            raise HTTPException(
                status_code=400,
                detail="File must be a CSV or Excel (.xlsx) file"
            )
        
        # Apply hypercube suppression
        protection_rules = ProtectionRules(
            min_frequency=min_frequency,
            dominance_n=dominance_n,
            dominance_k=dominance_k,
            p_percent=p_percent
        )
        suppressed_data, statistics = hypercube_suppress(
            data=df,
            protection_rules=protection_rules
        )
        
        # Get primary and secondary cell coordinates
        primary_coords = set()
        for cell in statistics.get('primary_cells', []):
            primary_coords.add((cell['row'], cell['col']))
        
        secondary_coords = set()
        for cell in statistics.get('secondary_cells', []):
            secondary_coords.add((cell['row'], cell['col']))
        
        # Generate downloadable file
        if output_format == "csv":
            stream = StringIO()
            suppressed_data.to_csv(stream, index=False)
            return StreamingResponse(
                iter([stream.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=suppressed_data.csv"}
            )
        elif output_format == "excel":
            # Export with color formatting
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            for c_idx, col_name in enumerate(suppressed_data.columns, 1):
                ws.cell(row=1, column=c_idx, value=str(col_name))
            
            # Write data with color formatting
            for r_idx in range(len(suppressed_data)):
                for c_idx in range(len(suppressed_data.columns)):
                    value = suppressed_data.iloc[r_idx, c_idx]
                    
                    # Convert to Excel-friendly format
                    if pd.isna(value):
                        cell_value = None
                    elif isinstance(value, (int, float, np.number)):
                        cell_value = float(value)
                    else:
                        cell_value = str(value)
                    
                    cell = ws.cell(row=r_idx+2, column=c_idx+1, value=cell_value)
                    
                    # Apply colors
                    if (r_idx, c_idx) in primary_coords:
                        cell.font = Font(color="0000FF", bold=True)  # Blue for primary
                    elif (r_idx, c_idx) in secondary_coords:
                        cell.font = Font(color="FF0000", bold=True)  # Red for secondary
            
            stream = BytesIO()
            wb.save(stream)
            excel_bytes = stream.getvalue()
            
            from fastapi.responses import Response
            return Response(
                content=excel_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=suppressed_data.xlsx"}
            )
        else:
            raise HTTPException(
                status_code=400,
                detail="output_format must be 'csv' or 'excel'"
            )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error processing table: {str(e)}"
        )


@app.post("/suppress/primary/")
async def primary_suppression(
    file: UploadFile = File(...),
    min_frequency: int = Form(3),
    dominance_n: int = Form(1),
    dominance_k: float = Form(80.0),
    p_percent: float = Form(10.0)
):
    """
    Apply Eurostat-compliant primary suppression to uploaded table
    
    Parameters:
    - file: CSV or Excel file containing the data table
    - min_frequency: Minimum frequency threshold (default: 3)
    - dominance_n: Number of top contributors for dominance rule (default: 1)
    - dominance_k: Percentage threshold for dominance (default: 80.0)
    - p_percent: P-percent rule threshold (default: 10.0)
    
    Returns:
    - JSON with primary suppressed table and statistics
    """
    import traceback
    import sys
    try:
        sys.stderr.write(f"\n[PRIMARY] Received file: {file.filename}\n")
        sys.stderr.write(f"[PRIMARY] Parameters: min_freq={min_frequency}, dom_n={dominance_n}, dom_k={dominance_k}, p={p_percent}\n")
        sys.stderr.flush()
        # Read uploaded file
        contents = await file.read()
        filename = file.filename.lower()
        
        if filename.endswith('.csv'):
            s = str(contents, 'utf-8')
            df = pd.read_csv(StringIO(s))
        elif filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(contents))
        else:
            raise HTTPException(
                status_code=400,
                detail="File must be a CSV or Excel (.xlsx) file"
            )
        
        # Detect value column (assume first numeric column or column named 'value')
        value_column = None
        if 'value' in df.columns:
            value_column = 'value'
        else:
            # Find first numeric column
            for col in df.columns:
                if pd.api.types.is_numeric_dtype(df[col]):
                    value_column = col
                    break
        
        if not value_column:
            raise HTTPException(
                status_code=400,
                detail="No numeric value column found. Please ensure your data has a 'value' column or at least one numeric column."
            )
        
        # Create protection rules from parameters
        protection_rules = PrimaryProtectionRules(
            min_frequency=min_frequency,
            dominance_n=dominance_n,
            dominance_k=dominance_k,
            p_percent=p_percent
        )
        
        # Apply primary suppression
        suppressed_data, summary = apply_primary_suppression_to_file(
            df=df,
            rules=protection_rules,
            value_column=value_column
        )
        
        # Convert suppressed data to JSON-compatible format
        suppressed_json = suppressed_data.replace({np.nan: None}).to_dict(orient="records")
        
        result = {
            "status": "success",
            "method": "primary_suppression",
            "summary": summary,
            "suppressed_data": suppressed_json,
            "column_names": suppressed_data.columns.tolist()
        }
        
        return result
        
    except Exception as e:
        sys.stderr.write(f"\n[PRIMARY ERROR] Exception occurred: {str(e)}\n")
        sys.stderr.flush()
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error processing table: {str(e)}"
        )


@app.post("/suppress/primary/download/")
async def primary_suppression_download(
    file: UploadFile = File(...),
    min_frequency: int = Form(3),
    dominance_n: int = Form(1),
    dominance_k: float = Form(80.0),
    p_percent: float = Form(10.0),
    output_format: str = Form("excel")
):
    """
    Apply primary suppression and download result as file with color highlighting
    
    Parameters:
    - file: CSV or Excel file containing the data table
    - min_frequency: Minimum frequency threshold (default: 3)
    - dominance_n: Number of top contributors for dominance rule (default: 1)
    - dominance_k: Percentage threshold for dominance (default: 80.0)
    - p_percent: P-percent rule threshold (default: 10.0)
    - output_format: Output format ('csv' or 'excel')
    
    Returns:
    - Downloadable file with primary suppressed data (red highlighting in Excel)
    """
    try:
        # Read uploaded file
        contents = await file.read()
        filename = file.filename.lower()
        
        if filename.endswith('.csv'):
            s = str(contents, 'utf-8')
            df = pd.read_csv(StringIO(s))
        elif filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(contents))
        else:
            raise HTTPException(
                status_code=400,
                detail="File must be a CSV or Excel (.xlsx) file"
            )
        
        # Detect value column
        value_column = None
        if 'value' in df.columns:
            value_column = 'value'
        else:
            for col in df.columns:
                if pd.api.types.is_numeric_dtype(df[col]):
                    value_column = col
                    break
        
        if not value_column:
            raise HTTPException(
                status_code=400,
                detail="No numeric value column found."
            )
        
        # Keep original data for display
        original_df = df.copy()
        
        # Apply primary suppression
        protection_rules = PrimaryProtectionRules(
            min_frequency=min_frequency,
            dominance_n=dominance_n,
            dominance_k=dominance_k,
            p_percent=p_percent
        )
        suppressed_data, summary = apply_primary_suppression_to_file(
            df=df,
            rules=protection_rules,
            value_column=value_column
        )
        
        # Get primary cell coordinates
        primary_coords = set()
        for cell in summary.get('primary_cells', []):
            primary_coords.add((cell['row'], cell['col']))
        
        # Generate downloadable file
        if output_format == "csv":
            stream = StringIO()
            suppressed_data.to_csv(stream, index=False)
            return StreamingResponse(
                iter([stream.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=primary_suppressed_data.csv"}
            )
        elif output_format == "excel":
            # Export with color formatting - show original values with red highlighting
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            for c_idx, col_name in enumerate(original_df.columns, 1):
                ws.cell(row=1, column=c_idx, value=str(col_name))
            
            # Write data with color formatting
            for r_idx in range(len(original_df)):
                for c_idx in range(len(original_df.columns)):
                    value = original_df.iloc[r_idx, c_idx]
                    
                    # Convert to Excel-friendly format
                    if pd.isna(value):
                        cell_value = None
                    elif isinstance(value, (int, float, np.number)):
                        cell_value = float(value)
                    else:
                        cell_value = str(value)
                    
                    cell = ws.cell(row=r_idx+2, column=c_idx+1, value=cell_value)
                    
                    # Highlight primary suppressed cells with red background
                    if (r_idx, c_idx) in primary_coords:
                        cell.font = Font(color="FFFFFF", bold=True)  # White text
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red background
            
            stream = BytesIO()
            wb.save(stream)
            excel_bytes = stream.getvalue()
            
            from fastapi.responses import Response
            return Response(
                content=excel_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=primary_suppressed_data.xlsx"}
            )
        else:
            raise HTTPException(
                status_code=400,
                detail="output_format must be 'csv' or 'excel'"
            )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error processing table: {str(e)}"
        )


@app.post("/batch/parse/")
async def parse_batch(
    batch_file: UploadFile = File(...)
):
    """
    Parse a τ-ARGUS batch file (.arb) and return the parsed commands.
    
    This endpoint parses the batch file and returns what commands it contains,
    without executing them. Useful for validation and preview.
    
    Parameters:
    - batch_file: .arb batch file
    
    Returns:
    - Parsed batch file structure with all commands
    """
    try:
        # Check file extension
        if not batch_file.filename.lower().endswith('.arb'):
            raise HTTPException(
                status_code=400,
                detail="File must be a .arb batch file"
            )
        
        # Read file contents
        contents = await batch_file.read()
        content_str = contents.decode('utf-8')
        
        # Parse batch file
        from .batch_parser import BatchParser
        batch = BatchParser.parse_content(content_str, base_path=None)
        
        # Convert to JSON-serializable format
        result = {
            "status": "success",
            "filename": batch_file.filename,
            "commands": [
                {
                    "command": cmd.command,
                    "parameters": cmd.parameters
                }
                for cmd in batch.commands
            ],
            "summary": {
                "microdata_file": batch.microdata_file,
                "table_data_file": batch.table_data_file,
                "metadata_file": batch.metadata_file,
                "table_spec": batch.table_spec,
                "safety_rules": [
                    {
                        "type": rule.rule_type,
                        "parameters": rule.parameters
                    }
                    for rule in batch.safety_rules
                ],
                "suppression_method": batch.suppression_method,
                "output_file": batch.output_file,
                "output_format": batch.output_format
            }
        }
        
        return result
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error parsing batch file: {str(e)}"
        )


@app.post("/batch/execute/")
async def execute_batch(
    batch_file: UploadFile = File(...),
    data_file: UploadFile = File(None),
    metadata_file: UploadFile = File(None),
    min_frequency: int = Form(10),
    dominance_n: int = Form(1),
    dominance_k: float = Form(80.0),
    p_percent: float = Form(10.0)
):
    """
    Execute a τ-ARGUS batch file.
    
    This endpoint parses and executes a batch file. You can optionally upload
    the data and metadata files along with the batch file, or the batch file
    can reference files that are already on the server.
    
    Parameters:
    - batch_file: .arb batch file
    - data_file: Optional .tab file (if not referenced in batch file)
    - metadata_file: Optional .rda file (if not referenced in batch file)
    
    Returns:
    - Execution results with suppressed data
    """
    temp_dir = None
    
    try:
        # Create temporary directory for uploaded files
        temp_dir = tempfile.mkdtemp()
        
        # Save batch file
        batch_path = Path(temp_dir) / batch_file.filename
        with open(batch_path, 'wb') as f:
            f.write(await batch_file.read())
        
        # Save optional data files
        if data_file:
            data_path = Path(temp_dir) / data_file.filename
            with open(data_path, 'wb') as f:
                f.write(await data_file.read())
        
        if metadata_file:
            meta_path = Path(temp_dir) / metadata_file.filename
            with open(meta_path, 'wb') as f:
                f.write(await metadata_file.read())
        
        # Parse batch file
        batch = parse_batch_file(str(batch_path))
        
        # Validate required files exist
        if not batch.table_data_file and not batch.microdata_file:
            raise HTTPException(
                status_code=400,
                detail="Batch file must specify a data file (OPENMICRODATA or OPENTABLEDATA)"
            )
        
        # Use uploaded files if provided, otherwise use paths from batch file
        if data_file:
            data_file_path = str(Path(temp_dir) / data_file.filename)
        else:
            data_file_path = batch.table_data_file or batch.microdata_file
            if not Path(data_file_path).exists():
                raise HTTPException(
                    status_code=400,
                    detail=f"Data file not found: {data_file_path}. Please upload it."
                )
        
        # Metadata is optional
        metadata_file_path = None
        if metadata_file:
            metadata_file_path = str(Path(temp_dir) / metadata_file.filename)
        elif batch.metadata_file and Path(batch.metadata_file).exists():
            metadata_file_path = batch.metadata_file
        
        # Parse data file using the resolved path
        if batch.table_data_file or data_file:
            # Tabulated data (.tab)
            df = TauArgusFormatHandler.parse_tab_file(data_file_path)
        else:
            # Microdata (.asc) - not yet implemented
            raise HTTPException(
                status_code=501,
                detail="Microdata (.asc) parsing not yet implemented. Please use tabulated data (.tab)."
            )
        
        # Parse metadata (optional)
        metadata = None
        if metadata_file_path:
            try:
                metadata = TauArgusFormatHandler.parse_rda_file(metadata_file_path)
            except Exception as e:
                # Metadata parsing failed - continue without it
                print(f"Warning: Could not parse metadata file: {e}")
        
        # Execute suppression (default to HYPERCUBE if not specified)
        method = batch.suppression_method or 'HYPERCUBE'
        
        if method in ['HYPERCUBE', 'GH', None]:
            # Use protection rules from UI (UI always wins)
            protection_rules = ProtectionRules(
                min_frequency=min_frequency,
                dominance_n=dominance_n,
                dominance_k=dominance_k,
                p_percent=p_percent
            )
            # Note: Batch file rules are ignored - UI rules take precedence
            
            # Apply hypercube suppression
            suppressed_data, statistics = hypercube_suppress(
                data=df,
                protection_rules=protection_rules
            )
            
            result = {
                "status": "success",
                "method": method or "hypercube",
                "batch_file": batch_file.filename,
                "data_file": Path(data_file_path).name,
                "metadata_file": Path(metadata_file_path).name if metadata_file_path else None,
                "safety_rules": [
                    {"type": rule.rule_type, "parameters": rule.parameters}
                    for rule in batch.safety_rules
                ],
                "statistics": statistics,
                "suppressed_data": suppressed_data.replace({np.nan: None}).to_dict(orient="records"),
                "column_names": suppressed_data.columns.tolist()
            }
            
            return result
        else:
            raise HTTPException(
                status_code=501,
                detail=f"Suppression method '{method}' not yet implemented. Only HYPERCUBE/GH supported."
            )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error executing batch file: {str(e)}"
        )
    finally:
        # Cleanup temporary directory
        if temp_dir and Path(temp_dir).exists():
            import shutil
            try:
                shutil.rmtree(temp_dir)
            except:
                pass


@app.post("/batch/download/")
async def download_batch_results(
    batch_file: UploadFile = File(...),
    data_file: UploadFile = File(...),
    metadata_file: UploadFile = File(None),
    min_frequency: int = Form(10),
    dominance_n: int = Form(1),
    dominance_k: float = Form(80.0),
    p_percent: float = Form(10.0),
    output_format: str = Form("excel")
):
    """
    Execute batch file and download results as Excel with color formatting.
    """
    temp_dir = None
    
    try:
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        
        # Save files
        batch_path = Path(temp_dir) / batch_file.filename
        with open(batch_path, 'wb') as f:
            f.write(await batch_file.read())
        
        data_path = Path(temp_dir) / data_file.filename
        with open(data_path, 'wb') as f:
            f.write(await data_file.read())
        
        if metadata_file:
            meta_path = Path(temp_dir) / metadata_file.filename
            with open(meta_path, 'wb') as f:
                f.write(await metadata_file.read())
        
        # Parse and execute
        batch = parse_batch_file(str(batch_path))
        df = TauArgusFormatHandler.parse_tab_file(str(data_path))
        
        # Apply suppression
        protection_rules = ProtectionRules(
            min_frequency=min_frequency,
            dominance_n=dominance_n,
            dominance_k=dominance_k,
            p_percent=p_percent
        )
        
        suppressed_data, statistics = hypercube_suppress(
            data=df,
            protection_rules=protection_rules
        )
        
        # Get cell coordinates
        primary_coords = set(
            (cell['row'], cell['col']) for cell in statistics.get('primary_cells', [])
        )
        secondary_coords = set(
            (cell['row'], cell['col']) for cell in statistics.get('secondary_cells', [])
        )
        
        # Generate Excel with colors
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        for c_idx, col_name in enumerate(suppressed_data.columns, 1):
            ws.cell(row=1, column=c_idx, value=str(col_name))
        
        # Write data with colors
        for r_idx in range(len(suppressed_data)):
            for c_idx in range(len(suppressed_data.columns)):
                value = suppressed_data.iloc[r_idx, c_idx]
                
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (int, float, np.number)):
                    cell_value = float(value)
                else:
                    cell_value = str(value)
                
                cell = ws.cell(row=r_idx+2, column=c_idx+1, value=cell_value)
                
                # Apply colors
                if (r_idx, c_idx) in primary_coords:
                    cell.font = Font(color="0000FF", bold=True)  # Blue
                elif (r_idx, c_idx) in secondary_coords:
                    cell.font = Font(color="FF0000", bold=True)  # Red
        
        stream = BytesIO()
        wb.save(stream)
        excel_bytes = stream.getvalue()
        
        from fastapi.responses import Response
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=batch_suppressed.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error downloading batch results: {str(e)}"
        )
    finally:
        if temp_dir and Path(temp_dir).exists():
            import shutil
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
